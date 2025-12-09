"""Data exporters for various formats.

Provides exporters for CSV, JSON, and Excel formats with streaming support.
"""

from __future__ import annotations

from abc import ABC, abstractmethod
import csv
from datetime import UTC, datetime
import io
import json
import logging
from typing import TYPE_CHECKING, Any, TypeVar
from uuid import UUID

if TYPE_CHECKING:
    from pathlib import Path

logger = logging.getLogger(__name__)

T = TypeVar("T")


def serialize_value(value: Any) -> Any:
    """Serialize a value to a JSON-compatible format.

    Args:
        value: Value to serialize.

    Returns:
        JSON-serializable value.
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, UUID):
        return str(value)
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    if hasattr(value, "model_dump"):
        return value.model_dump()
    if hasattr(value, "__dict__") and not isinstance(
        value, (str, int, float, bool, list, dict)
    ):
        return {
            k: serialize_value(v)
            for k, v in value.__dict__.items()
            if not k.startswith("_")
        }
    return value


class BaseExporter[T](ABC):
    """Base class for data exporters.

    Provides common functionality for exporting data to various formats.
    """

    def __init__(
        self,
        fields: list[str] | None = None,
        include_headers: bool = True,
    ) -> None:
        """Initialize exporter.

        Args:
            fields: Specific fields to export (all if None).
            include_headers: Whether to include headers in output.
        """
        self.fields = fields
        self.include_headers = include_headers

    @abstractmethod
    def export(self, records: list[T], output_path: Path) -> int:
        """Export records to a file.

        Args:
            records: Records to export.
            output_path: Path to output file.

        Returns:
            Number of records exported.
        """
        ...

    @abstractmethod
    def export_to_bytes(self, records: list[T]) -> bytes:
        """Export records to bytes (for streaming).

        Args:
            records: Records to export.

        Returns:
            Exported data as bytes.
        """
        ...

    @property
    @abstractmethod
    def content_type(self) -> str:
        """MIME content type for this format."""
        ...

    @property
    @abstractmethod
    def file_extension(self) -> str:
        """File extension for this format."""
        ...

    def _extract_fields(self, record: T) -> dict[str, Any]:
        """Extract field values from a record.

        Args:
            record: Record to extract from.

        Returns:
            Dictionary of field values.
        """
        if hasattr(record, "model_dump"):
            data = record.model_dump()
        elif hasattr(record, "__dict__"):
            data = {k: v for k, v in record.__dict__.items() if not k.startswith("_")}
        else:
            data = dict(record) if hasattr(record, "__iter__") else {}

        # Filter to selected fields if specified
        if self.fields:
            data = {k: v for k, v in data.items() if k in self.fields}

        # Serialize values
        return {k: serialize_value(v) for k, v in data.items()}


class CSVExporter(BaseExporter[T]):
    """Export data to CSV format."""

    @property
    def content_type(self) -> str:
        return "text/csv"

    @property
    def file_extension(self) -> str:
        return "csv"

    def export(self, records: list[T], output_path: Path) -> int:
        """Export records to CSV file."""
        if not records:
            output_path.write_text("")
            return 0

        rows = [self._extract_fields(r) for r in records]
        fieldnames = list(rows[0].keys()) if rows else []

        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            if self.include_headers:
                writer.writeheader()
            writer.writerows(rows)

        return len(rows)

    def export_to_bytes(self, records: list[T]) -> bytes:
        """Export records to CSV bytes."""
        if not records:
            return b""

        rows = [self._extract_fields(r) for r in records]
        fieldnames = list(rows[0].keys()) if rows else []

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        if self.include_headers:
            writer.writeheader()
        writer.writerows(rows)

        return output.getvalue().encode("utf-8")


class JSONExporter(BaseExporter[T]):
    """Export data to JSON format."""

    def __init__(
        self,
        fields: list[str] | None = None,
        include_headers: bool = True,
        indent: int = 2,
    ) -> None:
        """Initialize JSON exporter.

        Args:
            fields: Specific fields to export.
            include_headers: Not used for JSON.
            indent: JSON indentation level.
        """
        super().__init__(fields, include_headers)
        self.indent = indent

    @property
    def content_type(self) -> str:
        return "application/json"

    @property
    def file_extension(self) -> str:
        return "json"

    def export(self, records: list[T], output_path: Path) -> int:
        """Export records to JSON file."""
        rows = [self._extract_fields(r) for r in records]

        export_data = {
            "exported_at": datetime.now(UTC).isoformat(),
            "record_count": len(rows),
            "records": rows,
        }

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(
                export_data, f, indent=self.indent, ensure_ascii=False, default=str
            )

        return len(rows)

    def export_to_bytes(self, records: list[T]) -> bytes:
        """Export records to JSON bytes."""
        rows = [self._extract_fields(r) for r in records]

        export_data = {
            "exported_at": datetime.now(UTC).isoformat(),
            "record_count": len(rows),
            "records": rows,
        }

        return json.dumps(
            export_data, indent=self.indent, ensure_ascii=False, default=str
        ).encode("utf-8")


class ExcelExporter(BaseExporter[T]):
    """Export data to Excel format.

    Requires openpyxl to be installed.
    """

    @property
    def content_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    @property
    def file_extension(self) -> str:
        return "xlsx"

    def _build_workbook(self, records: list[T]) -> tuple[Any, int]:
        """Build an Excel workbook from records.

        Args:
            records: Records to export.

        Returns:
            Tuple of (workbook, record_count).
        """
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except ImportError as err:
            msg = "openpyxl is required for Excel export. Install with: pip install openpyxl"
            raise ImportError(msg) from err

        wb = Workbook()
        ws = wb.active
        if ws is None:
            msg = "Workbook has no active worksheet"
            raise RuntimeError(msg)
        ws.title = "Export"

        if not records:
            return wb, 0

        rows = [self._extract_fields(r) for r in records]
        fieldnames = list(rows[0].keys()) if rows else []

        # Write headers
        if self.include_headers:
            for col, header in enumerate(fieldnames, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = cell.font.copy(bold=True)

        # Write data
        start_row = 2 if self.include_headers else 1
        for row_idx, row_data in enumerate(rows, start_row):
            for col, field in enumerate(fieldnames, 1):
                value = row_data.get(field)
                # Excel can't handle some types directly
                if isinstance(value, (dict, list)):
                    value = json.dumps(value)
                ws.cell(row=row_idx, column=col, value=value)

        # Auto-size columns
        for col in range(1, len(fieldnames) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].auto_size = True

        return wb, len(rows)

    def export(self, records: list[T], output_path: Path) -> int:
        """Export records to Excel file."""
        wb, record_count = self._build_workbook(records)
        wb.save(output_path)
        return record_count

    def export_to_bytes(self, records: list[T]) -> bytes:
        """Export records to Excel bytes."""
        wb, _ = self._build_workbook(records)
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


def get_exporter(
    format: str,
    fields: list[str] | None = None,
    include_headers: bool = True,
) -> BaseExporter:
    """Get the appropriate exporter for a format.

    Args:
        format: Export format (csv, json, xlsx).
        fields: Specific fields to export.
        include_headers: Whether to include headers.

    Returns:
        Exporter instance.

    Raises:
        ValueError: If format is not supported.
    """
    format_lower = format.lower()
    if format_lower == "csv":
        return CSVExporter(fields=fields, include_headers=include_headers)
    if format_lower == "json":
        return JSONExporter(fields=fields, include_headers=include_headers)
    if format_lower in ("xlsx", "excel"):
        return ExcelExporter(fields=fields, include_headers=include_headers)
    raise ValueError(f"Unsupported export format: {format}")


__all__ = [
    "BaseExporter",
    "CSVExporter",
    "ExcelExporter",
    "JSONExporter",
    "get_exporter",
    "serialize_value",
]
