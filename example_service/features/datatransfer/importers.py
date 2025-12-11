"""Data importers for various formats.

Provides importers for CSV, JSON, and Excel formats with validation.
"""

from __future__ import annotations

from abc import ABC, abstractmethod
import csv
from datetime import datetime
import io
import json
import logging
from typing import TYPE_CHECKING, Any, TypeVar
from uuid import UUID

if TYPE_CHECKING:
    from collections.abc import Callable
    from pathlib import Path

logger = logging.getLogger(__name__)

T = TypeVar("T")


class DataImportError(Exception):
    """Error during import operation.

    Named DataImportError to avoid shadowing Python's built-in ImportError.
    """

    def __init__(self, message: str, row: int | None = None, field: str | None = None) -> None:
        super().__init__(message)
        self.row = row
        self.field = field


class ParsedRecord:
    """A parsed record from import file."""

    def __init__(
        self,
        row_number: int,
        data: dict[str, Any],
        errors: list[tuple[str, str]] | None = None,
    ) -> None:
        """Initialize parsed record.

        Args:
            row_number: 1-indexed row number in source file.
            data: Parsed record data.
            errors: List of (field, error_message) tuples.
        """
        self.row_number = row_number
        self.data = data
        self.errors = errors or []

    @property
    def is_valid(self) -> bool:
        """Whether the record has no validation errors."""
        return len(self.errors) == 0


class BaseImporter[T](ABC):
    """Base class for data importers.

    Provides common functionality for importing data from various formats.
    """

    def __init__(
        self,
        required_fields: list[str] | None = None,
        field_types: dict[str, type] | None = None,
        field_validators: dict[str, Callable[[Any], bool]] | None = None,
    ) -> None:
        """Initialize importer.

        Args:
            required_fields: Fields that must be present.
            field_types: Expected type for each field.
            field_validators: Custom validator functions for fields.
        """
        self.required_fields = required_fields or []
        self.field_types = field_types or {}
        self.field_validators = field_validators or {}

    @abstractmethod
    def parse_file(self, file_path: Path) -> list[ParsedRecord]:
        """Parse records from a file.

        Args:
            file_path: Path to input file.

        Returns:
            List of parsed records with validation results.
        """
        ...

    @abstractmethod
    def parse_bytes(self, data: bytes) -> list[ParsedRecord]:
        """Parse records from bytes.

        Args:
            data: Input data as bytes.

        Returns:
            List of parsed records with validation results.
        """
        ...

    @property
    @abstractmethod
    def supported_extensions(self) -> list[str]:
        """Supported file extensions for this importer."""
        ...

    def validate_record(self, record: dict[str, Any], *, row_number: int) -> ParsedRecord:
        """Validate a parsed record.

        Args:
            row_number: Row number for error reporting.
            record: Raw record data.

        Returns:
            ParsedRecord with validation results.
        """
        errors: list[tuple[str, str]] = []
        validated_data = {}

        # Check required fields
        errors.extend(
            (field, f"Required field '{field}' is missing or empty")
            for field in self.required_fields
            if field not in record or record[field] in (None, "")
        )

        # Type conversion and validation
        for field, value in record.items():
            field_value = value
            try:
                # Type conversion
                if (
                    field in self.field_types
                    and field_value is not None
                    and field_value != ""
                ):
                    expected_type = self.field_types[field]
                    try:
                        if expected_type is bool:
                            if isinstance(field_value, str):
                                converted = field_value.lower() in (
                                    "true",
                                    "1",
                                    "yes",
                                    "on",
                                )
                            else:
                                converted = bool(field_value)
                            field_value = converted
                        elif expected_type is int:
                            field_value = int(field_value)
                        elif expected_type is float:
                            field_value = float(field_value)
                        elif expected_type is datetime:
                            if isinstance(field_value, datetime):
                                pass  # Already a datetime
                            elif isinstance(field_value, str):
                                # Try ISO format first, then common formats
                                for fmt in (
                                    None,  # ISO format via fromisoformat
                                    "%Y-%m-%d %H:%M:%S",
                                    "%Y-%m-%d",
                                    "%d/%m/%Y",
                                    "%m/%d/%Y",
                                ):
                                    try:
                                        if fmt is None:
                                            field_value = datetime.fromisoformat(
                                                field_value,
                                            )
                                        else:
                                            field_value = datetime.strptime(
                                                field_value,
                                                fmt,
                                            )
                                        break
                                    except ValueError:
                                        continue
                                else:
                                    msg = (
                                        "Unrecognized datetime format:"
                                        f" {field_value}"
                                    )
                                    raise ValueError(msg)
                            else:
                                msg = (
                                    "Cannot convert"
                                    f" {type(field_value).__name__} to datetime"
                                )
                                raise TypeError(msg)
                        elif expected_type is UUID:
                            if isinstance(field_value, UUID):
                                pass  # Already a UUID
                            elif isinstance(field_value, str):
                                field_value = UUID(field_value)
                            else:
                                msg = (
                                    "Cannot convert"
                                    f" {type(field_value).__name__} to UUID"
                                )
                                raise TypeError(msg)
                        else:
                            field_value = expected_type(field_value)
                    except (ValueError, TypeError) as e:
                        errors.append(
                            (
                                field,
                                f"Cannot convert '{field_value}' to"
                                f" {expected_type.__name__}: {e}",
                            ),
                        )
                        continue

                # Custom validation
                if field in self.field_validators:
                    validator = self.field_validators[field]
                    try:
                        validation_result = validator(field_value)
                        if validation_result is not True and validation_result is not None:
                            errors.append((field, str(validation_result)))
                    except Exception as e:
                        errors.append((field, f"Validation failed: {e}"))

                validated_data[field] = field_value

            except Exception as e:
                errors.append((field, f"Processing error: {e}"))

        return ParsedRecord(row_number=row_number, data=validated_data, errors=errors)


class CSVImporter(BaseImporter[T]):
    """Import data from CSV format."""

    def __init__(
        self,
        required_fields: list[str] | None = None,
        field_types: dict[str, type] | None = None,
        field_validators: dict[str, Callable[[Any], bool]] | None = None,
        delimiter: str = ",",
        encoding: str = "utf-8",
    ) -> None:
        """Initialize CSV importer.

        Args:
            required_fields: Required fields.
            field_types: Field type mappings.
            field_validators: Custom validators.
            delimiter: CSV delimiter character.
            encoding: File encoding.
        """
        super().__init__(required_fields, field_types, field_validators)
        self.delimiter = delimiter
        self.encoding = encoding

    @property
    def supported_extensions(self) -> list[str]:
        """Return supported file extensions for CSV imports."""
        return [".csv"]

    def parse_file(self, file_path: Path) -> list[ParsedRecord]:
        """Parse CSV file."""
        with open(file_path, encoding=self.encoding, newline="") as f:
            return self._parse_csv(f)

    def parse_bytes(self, data: bytes) -> list[ParsedRecord]:
        """Parse CSV bytes."""
        text = data.decode(self.encoding)
        return self._parse_csv(io.StringIO(text))

    def _parse_csv(self, file_obj: io.StringIO | Any) -> list[ParsedRecord]:
        """Parse CSV from file-like object."""
        reader = csv.DictReader(file_obj, delimiter=self.delimiter)
        records = []

        for row_num, row in enumerate(reader, start=2):  # Row 1 is header
            # Clean up None values from empty strings
            cleaned = {k: (v if v != "" else None) for k, v in row.items() if k}
            parsed = self.validate_record(cleaned, row_number=row_num)
            records.append(parsed)

        return records


class JSONImporter(BaseImporter[T]):
    """Import data from JSON format."""

    def __init__(
        self,
        required_fields: list[str] | None = None,
        field_types: dict[str, type] | None = None,
        field_validators: dict[str, Callable[[Any], bool]] | None = None,
        records_key: str = "records",
    ) -> None:
        """Initialize JSON importer.

        Args:
            required_fields: Required fields.
            field_types: Field type mappings.
            field_validators: Custom validators.
            records_key: Key containing records array in JSON.
        """
        super().__init__(required_fields, field_types, field_validators)
        self.records_key = records_key

    @property
    def supported_extensions(self) -> list[str]:
        """Return supported file extensions for JSON imports."""
        return [".json"]

    def parse_file(self, file_path: Path) -> list[ParsedRecord]:
        """Parse JSON file."""
        with open(file_path, encoding="utf-8") as f:
            data = json.load(f)
        return self._parse_json(data)

    def parse_bytes(self, data: bytes) -> list[ParsedRecord]:
        """Parse JSON bytes."""
        parsed = json.loads(data.decode("utf-8"))
        return self._parse_json(parsed)

    def _parse_json(self, data: dict | list) -> list[ParsedRecord]:
        """Parse JSON data structure."""
        # Support both array and object with records key
        if isinstance(data, list):
            records_list = data
        elif isinstance(data, dict):
            records_list = data.get(self.records_key, [])
            if not isinstance(records_list, list):
                msg = f"Expected '{self.records_key}' to be an array"
                raise DataImportError(msg)
        else:
            msg = "JSON must be an array or object with records"
            raise DataImportError(msg)

        records = []
        for row_num, row in enumerate(records_list, start=1):
            if not isinstance(row, dict):
                records.append(
                    ParsedRecord(
                        row_number=row_num,
                        data={},
                        errors=[("_record", f"Row {row_num} is not an object")],
                    ),
                )
                continue

            parsed = self.validate_record(row, row_number=row_num)
            records.append(parsed)

        return records


class ExcelImporter(BaseImporter[T]):
    """Import data from Excel format.

    Requires openpyxl to be installed.
    """

    def __init__(
        self,
        required_fields: list[str] | None = None,
        field_types: dict[str, type] | None = None,
        field_validators: dict[str, Callable[[Any], bool]] | None = None,
        sheet_name: str | None = None,
    ) -> None:
        """Initialize Excel importer.

        Args:
            required_fields: Required fields.
            field_types: Field type mappings.
            field_validators: Custom validators.
            sheet_name: Specific sheet to import (first sheet if None).
        """
        super().__init__(required_fields, field_types, field_validators)
        self.sheet_name = sheet_name

    @property
    def supported_extensions(self) -> list[str]:
        """Return supported file extensions for Excel imports."""
        return [".xlsx", ".xls"]

    def parse_file(self, file_path: Path) -> list[ParsedRecord]:
        """Parse Excel file."""
        try:
            from openpyxl import load_workbook
        except ImportError as err:
            msg = "openpyxl is required for Excel import"
            raise DataImportError(msg) from err

        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[self.sheet_name] if self.sheet_name else wb.active

        return self._parse_worksheet(ws)

    def parse_bytes(self, data: bytes) -> list[ParsedRecord]:
        """Parse Excel bytes."""
        try:
            from openpyxl import load_workbook
        except ImportError as err:
            msg = "openpyxl is required for Excel import"
            raise DataImportError(msg) from err

        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb[self.sheet_name] if self.sheet_name else wb.active

        return self._parse_worksheet(ws)

    def _parse_worksheet(self, ws: Any) -> list[ParsedRecord]:
        """Parse worksheet rows."""
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        # First row is headers
        headers = [str(h).strip() if h else f"column_{i}" for i, h in enumerate(rows[0])]

        records = []
        for row_num, row_data in enumerate(rows[1:], start=2):
            # Create dict from row
            row_dict = {}
            for i, value in enumerate(row_data):
                if i < len(headers):
                    row_dict[headers[i]] = value

            parsed = self.validate_record(row_dict, row_number=row_num)
            records.append(parsed)

        return records


def get_importer(
    format: str,
    required_fields: list[str] | None = None,
    field_types: dict[str, type] | None = None,
    field_validators: dict[str, Callable[[Any], bool]] | None = None,
) -> BaseImporter:
    """Get the appropriate importer for a format.

    Args:
        format: Import format (csv, json, xlsx).
        required_fields: Required fields.
        field_types: Field type mappings.
        field_validators: Custom validators.

    Returns:
        Importer instance.

    Raises:
        ValueError: If format is not supported.
    """
    format_lower = format.lower()
    if format_lower == "csv":
        return CSVImporter(
            required_fields=required_fields,
            field_types=field_types,
            field_validators=field_validators,
        )
    if format_lower == "json":
        return JSONImporter(
            required_fields=required_fields,
            field_types=field_types,
            field_validators=field_validators,
        )
    if format_lower in ("xlsx", "excel"):
        return ExcelImporter(
            required_fields=required_fields,
            field_types=field_types,
            field_validators=field_validators,
        )
    msg = f"Unsupported import format: {format}"
    raise ValueError(msg)


__all__ = [
    "BaseImporter",
    "CSVImporter",
    "DataImportError",
    "ExcelImporter",
    "JSONImporter",
    "ParsedRecord",
    "get_importer",
]
